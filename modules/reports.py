import io
import re

import numpy as np
import pandas as pd
import streamlit as st

from .classification import ajustar_macro_iva_105, clasificar
from .formatting import fmt_ar, metric_full
from .parsing import find_macro_dyc_total_from_lines, find_saldo_anterior_from_lines, find_saldo_final_from_lines, parse_lines

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    REPORTLAB_OK = True
except Exception:
    REPORTLAB_OK = False


def _account_suffix(account_number: str) -> str:
    return f"_{re.sub(r'[^0-9A-Za-z]+', '_', account_number)}"


def _build_movements_dataframe(lines: list[str]) -> tuple[pd.DataFrame, pd.Timestamp, float, float]:
    df = parse_lines(lines)
    fecha_cierre, saldo_final_pdf = find_saldo_final_from_lines(lines)
    saldo_anterior = find_saldo_anterior_from_lines(lines)

    if df.empty:
        return df, fecha_cierre, saldo_final_pdf, saldo_anterior

    if not np.isnan(saldo_anterior):
        first_date = df["fecha"].dropna().min()
        fecha_apertura = (first_date - pd.Timedelta(days=1)).normalize() + pd.Timedelta(hours=23, minutes=59, seconds=59) if pd.notna(first_date) else pd.NaT
        apertura = pd.DataFrame([{
            "fecha": fecha_apertura,
            "descripcion": "SALDO ANTERIOR",
            "desc_norm": "SALDO ANTERIOR",
            "debito": 0.0,
            "credito": 0.0,
            "importe": 0.0,
            "saldo": float(saldo_anterior),
            "pagina": 0,
            "orden": 0,
        }])
        df = pd.concat([apertura, df], ignore_index=True)

    df = df.sort_values(["fecha", "orden"]).reset_index(drop=True)
    df["delta_saldo"] = df["saldo"].diff()
    df["debito"] = np.where(df["delta_saldo"] < 0, -df["delta_saldo"], 0.0)
    df["credito"] = np.where(df["delta_saldo"] > 0, df["delta_saldo"], 0.0)
    df["importe"] = df["debito"] - df["credito"]
    df["Clasificación"] = df.apply(
        lambda r: clasificar(str(r.get("descripcion", "")), str(r.get("desc_norm", "")), r.get("debito", 0.0), r.get("credito", 0.0)),
        axis=1,
    )
    df = ajustar_macro_iva_105(df)
    return df, fecha_cierre, saldo_final_pdf, saldo_anterior


def _download_excel(df: pd.DataFrame, label: str, file_name: str, key: str):
    def build_with_xlsxwriter() -> bytes:
        import xlsxwriter  # noqa: F401
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="Movimientos")
            wb = writer.book
            ws = writer.sheets["Movimientos"]
            money_fmt = wb.add_format({"num_format": "#,##0.00"})
            date_fmt = wb.add_format({"num_format": "dd/mm/yyyy"})
            for idx, col in enumerate(df.columns, start=0):
                col_values = df[col].astype(str)
                max_len = max(len(col), *(len(v) for v in col_values))
                ws.set_column(idx, idx, min(max_len + 2, 45))
            for c in ["debito", "credito", "importe", "saldo", "delta_saldo"]:
                if c in df.columns:
                    j = df.columns.get_loc(c)
                    ws.set_column(j, j, 16, money_fmt)
            if "fecha" in df.columns:
                j = df.columns.get_loc("fecha")
                ws.set_column(j, j, 14, date_fmt)
        return output.getvalue()

    def build_with_openpyxl() -> bytes:
        import openpyxl  # noqa: F401
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Movimientos")
            ws = writer.sheets["Movimientos"]
            for col_idx, col_name in enumerate(df.columns, start=1):
                max_len = max(len(str(col_name)), *(len(str(v)) for v in df[col_name].astype(str)))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 45)
                if col_name in ["debito", "credito", "importe", "saldo", "delta_saldo"]:
                    for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                        for c in cell:
                            c.number_format = '#,##0.00'
                if col_name == "fecha":
                    for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                        for c in cell:
                            c.number_format = 'DD/MM/YYYY'
        return output.getvalue()

    try:
        data = build_with_xlsxwriter()
    except Exception:
        try:
            data = build_with_openpyxl()
        except Exception as e:
            st.error(f"No se pudo generar el Excel. Revisá requirements.txt: xlsxwriter u openpyxl. Error: {e}")
            return

    st.download_button(
        label,
        data=data,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=key,
    )


def render_account_report(account_title: str, account_number: str, acc_id: str, lines: list[str]):
    st.markdown("---")
    st.subheader(f"{account_title} · Nro {account_number}")

    df, fecha_cierre, saldo_final_pdf, saldo_anterior = _build_movements_dataframe(lines)

    if df.empty:
        total_debitos = 0.0
        total_creditos = 0.0
        saldo_inicial = float(saldo_anterior) if not np.isnan(saldo_anterior) else 0.0
        saldo_final_visto = float(saldo_final_pdf) if not np.isnan(saldo_final_pdf) else saldo_inicial
        saldo_final_calculado = saldo_inicial + total_creditos - total_debitos
        diferencia = saldo_final_calculado - saldo_final_visto
        cuadra = abs(diferencia) < 0.01
        st.caption("Resumen del período")
        c1, c2, c3 = st.columns(3)
        with c1: metric_full("Saldo inicial", f"$ {fmt_ar(saldo_inicial)}")
        with c2: metric_full("Total créditos (+)", f"$ {fmt_ar(total_creditos)}")
        with c3: metric_full("Total débitos (–)", f"$ {fmt_ar(total_debitos)}")
        c4, c5, c6 = st.columns(3)
        with c4: metric_full("Saldo final (PDF)", f"$ {fmt_ar(saldo_final_visto)}")
        with c5: metric_full("Saldo final calculado", f"$ {fmt_ar(saldo_final_calculado)}")
        with c6: metric_full("Diferencia", f"$ {fmt_ar(diferencia)}")
        st.success("Conciliado.") if cuadra else st.error("No cuadra la conciliación.")
        if pd.notna(fecha_cierre):
            st.caption(f"Cierre según PDF: {fecha_cierre.strftime('%d/%m/%Y')}")
        st.info("Sin movimientos")
        return

    df_sorted = df.drop(columns=["orden"]).reset_index(drop=True)
    saldo_inicial = float(df_sorted.loc[0, "saldo"])
    total_debitos = float(df_sorted["debito"].sum())
    total_creditos = float(df_sorted["credito"].sum())
    saldo_final_visto = float(df_sorted["saldo"].iloc[-1]) if np.isnan(saldo_final_pdf) else float(saldo_final_pdf)
    saldo_final_calculado = saldo_inicial + total_creditos - total_debitos
    diferencia = saldo_final_calculado - saldo_final_visto
    cuadra = abs(diferencia) < 0.01

    date_suffix = f"_{fecha_cierre.strftime('%Y%m%d')}" if pd.notna(fecha_cierre) else ""
    acc_suffix = _account_suffix(account_number)

    st.caption("Resumen del período")
    c1, c2, c3 = st.columns(3)
    with c1: metric_full("Saldo inicial", f"$ {fmt_ar(saldo_inicial)}")
    with c2: metric_full("Total créditos (+)", f"$ {fmt_ar(total_creditos)}")
    with c3: metric_full("Total débitos (–)", f"$ {fmt_ar(total_debitos)}")
    c4, c5, c6 = st.columns(3)
    with c4: metric_full("Saldo final (PDF)", f"$ {fmt_ar(saldo_final_visto)}")
    with c5: metric_full("Saldo final calculado", f"$ {fmt_ar(saldo_final_calculado)}")
    with c6: metric_full("Diferencia", f"$ {fmt_ar(diferencia)}")
    st.success("Conciliado.") if cuadra else st.error("No cuadra la conciliación.")
    if pd.notna(fecha_cierre):
        st.caption(f"Cierre según PDF: {fecha_cierre.strftime('%d/%m/%Y')}")

    st.caption("Resumen Operativo: Registración Módulo IVA")
    iva21_mask = df_sorted["Clasificación"].eq("IVA 21% (sobre comisiones)")
    iva105_mask = df_sorted["Clasificación"].eq("IVA 10,5% (sobre comisiones)")
    iva21 = float(df_sorted.loc[iva21_mask, "debito"].sum())
    iva105 = float(df_sorted.loc[iva105_mask, "debito"].sum())
    net21 = round(iva21 / 0.21, 2) if iva21 else 0.0
    net105 = round(iva105 / 0.105, 2) if iva105 else 0.0
    percep_iva = float(df_sorted.loc[df_sorted["Clasificación"].eq("Percepciones de IVA"), "debito"].sum())
    ley_25413_mask = df_sorted["Clasificación"].eq("LEY 25.413")
    ley_25413_debitos = float(df_sorted.loc[ley_25413_mask, "debito"].sum())
    ley_25413_creditos = float(df_sorted.loc[ley_25413_mask, "credito"].sum())
    ley_25413 = ley_25413_debitos - ley_25413_creditos
    sircreb = float(df_sorted.loc[df_sorted["Clasificación"].eq("SIRCREB"), "debito"].sum())

    m1, m2, m3 = st.columns(3)
    with m1: metric_full("Neto Comisiones 21%", f"$ {fmt_ar(net21)}")
    with m2: metric_full("IVA 21%", f"$ {fmt_ar(iva21)}")
    with m3: metric_full("Bruto 21%", f"$ {fmt_ar(net21 + iva21)}")
    n1, n2, n3 = st.columns(3)
    with n1: metric_full("Neto Comisiones 10,5%", f"$ {fmt_ar(net105)}")
    with n2: metric_full("IVA 10,5%", f"$ {fmt_ar(iva105)}")
    with n3: metric_full("Bruto 10,5%", f"$ {fmt_ar(net105 + iva105)}")
    o1, o2, o3 = st.columns(3)
    with o1: metric_full("Percepciones de IVA", f"$ {fmt_ar(percep_iva)}")
    with o2: metric_full("Ley 25.413 / DyC", f"$ {fmt_ar(ley_25413)}")
    with o3: metric_full("SIRCREB", f"$ {fmt_ar(sircreb)}")

    dyc_pdf = find_macro_dyc_total_from_lines(lines)
    if not np.isnan(dyc_pdf):
        st.caption("Control Ley 25.413 / DyC contra total informado por Banco Macro")
        diferencia_dyc = ley_25413 - float(dyc_pdf)
        d1, d2, d3 = st.columns(3)
        with d1: metric_full("DyC calculado", f"$ {fmt_ar(ley_25413)}")
        with d2: metric_full("DyC informado PDF", f"$ {fmt_ar(dyc_pdf)}")
        with d3: metric_full("Diferencia", f"$ {fmt_ar(diferencia_dyc)}")
        st.success("Control DyC conciliado.") if abs(diferencia_dyc) < 0.01 else st.warning("Control DyC no coincide con el total informado por el PDF.")

    st.caption("Detalle de movimientos")
    df_view = df_sorted.copy()
    for c in ["debito", "credito", "importe", "saldo", "delta_saldo"]:
        if c in df_view.columns:
            df_view[c] = df_view[c].map(fmt_ar)
    st.dataframe(df_view, use_container_width=True)

    st.caption("Detalle de créditos (préstamos)")
    credit_classes = ["Cuota de préstamo", "Acreditación Préstamos"]
    df_creditos = df_sorted.loc[df_sorted["Clasificación"].isin(credit_classes)].copy()
    if df_creditos.empty:
        st.info("Sin movimientos de créditos/préstamos en el período.")
    else:
        total_cuotas = float(df_creditos.loc[df_creditos["Clasificación"].eq("Cuota de préstamo"), "debito"].sum())
        total_acredit = float(df_creditos.loc[df_creditos["Clasificación"].eq("Acreditación Préstamos"), "credito"].sum())
        neto_creditos = total_acredit - total_cuotas
        k1, k2, k3 = st.columns(3)
        with k1: metric_full("Acreditaciones de préstamos (+)", f"$ {fmt_ar(total_acredit)}")
        with k2: metric_full("Cuotas de préstamo (–)", f"$ {fmt_ar(total_cuotas)}")
        with k3: metric_full("Neto (acreditado – cuotas)", f"$ {fmt_ar(neto_creditos)}")
        df_creditos_view = df_creditos.copy()
        for c in ["debito", "credito", "importe", "saldo", "delta_saldo"]:
            if c in df_creditos_view.columns:
                df_creditos_view[c] = df_creditos_view[c].map(fmt_ar)
        st.dataframe(df_creditos_view, use_container_width=True)
        _download_excel(
            df_creditos,
            "📥 Descargar Excel – Detalle Créditos",
            f"detalle_creditos_macro{acc_suffix}{date_suffix}.xlsx",
            f"dl_creditos_xlsx_{acc_id}",
        )

    st.caption("Descargar")
    _download_excel(df_sorted, "📥 Descargar Excel", f"resumen_bancario_macro{acc_suffix}{date_suffix}.xlsx", f"dl_xlsx_{acc_id}")

    if REPORTLAB_OK:
        try:
            pdf_buf = io.BytesIO()
            doc = SimpleDocTemplate(pdf_buf, pagesize=A4, title="Resumen Operativo - Registración Módulo IVA")
            styles = getSampleStyleSheet()
            elems = [Paragraph("Resumen Operativo: Registración Módulo IVA", styles["Title"]), Spacer(1, 8)]
            datos = [
                ["Concepto", "Importe"],
                ["Neto Comisiones 21%", fmt_ar(net21)],
                ["IVA 21%", fmt_ar(iva21)],
                ["Bruto 21%", fmt_ar(net21 + iva21)],
                ["Neto Comisiones 10,5%", fmt_ar(net105)],
                ["IVA 10,5%", fmt_ar(iva105)],
                ["Bruto 10,5%", fmt_ar(net105 + iva105)],
                ["Percepciones de IVA (RG 3337 / RG 2408)", fmt_ar(percep_iva)],
                ["Ley 25.413 / DyC", fmt_ar(ley_25413)],
                ["SIRCREB", fmt_ar(sircreb)],
            ]
            if not np.isnan(dyc_pdf):
                datos.extend([
                    ["Control DyC informado PDF", fmt_ar(dyc_pdf)],
                    ["Diferencia DyC", fmt_ar(ley_25413 - float(dyc_pdf))],
                ])
            datos.append(["TOTAL", fmt_ar(net21 + iva21 + net105 + iva105 + percep_iva + ley_25413 + sircreb)])
            tbl = Table(datos, colWidths=[300, 120])
            tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ]))
            elems.append(tbl)
            elems.append(Spacer(1, 12))
            elems.append(Paragraph("Herramienta para uso interno - AIE San Justo", styles["Normal"]))
            doc.build(elems)
            st.download_button(
                "📄 Descargar PDF – Resumen Operativo (IVA)",
                data=pdf_buf.getvalue(),
                file_name=f"Resumen_Operativo_IVA_macro{acc_suffix}{date_suffix}.pdf",
                mime="application/pdf",
                use_container_width=True,
                key=f"dl_pdf_{acc_id}",
            )
        except Exception as e:
            st.info(f"No se pudo generar el PDF del Resumen Operativo: {e}")
